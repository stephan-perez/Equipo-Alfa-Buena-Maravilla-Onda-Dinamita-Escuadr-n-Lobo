# FUNDAMENTOS DE PROGRAMACIÓN PARA INGENIERÍA
# SECCIÓN DEL CURSO: L-7
# PROFESOR DE LABORATORIO: RICARDO CORBINAUD PÉREZ
# AYUDANTE DE LABORATORIO: FELIPE BAEZA MUÑOZ
# GRUPO: 6
# INTEGRANTES 
# 1. Vanitza Contreras Fernández , 
# 2. Vicente Guarda Antúnez , 
# 3. Martín Oliva Flores , 21.911.705-1
# 4. Stefan Pérez Quintana ,
# 5. May Pino López , 
# DESCRIPCIÓN DEL PROGRAMA ...<CONTINÚE CON EL PROGRAMA A PARTIR DE AQUÍ> 

#####################
#Bloque definiciones#
#####################

#Importaciones
import sys
import tkinter as tk
from tkinter import ttk
from PIL import ImageTk, Image
import openpyxl as ox
import matplotlib.pyplot as plt
import statistics as stats
from tkinter import Toplevel
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

#Constantes
#Datos obtenidos en excel
datos= []
hoja = ox.load_workbook("datos_secciones.xlsx") #Se abre el archivo
hoja_data = hoja.active
n_datos = 0
for fila in range(1, hoja_data.max_row): #Se agrega a la variable datos, los datos obtenidos en una lista de listas
    lista_fila = []
    n_datos += 1
    for columna in hoja_data.iter_cols(0, 6): #6 = hoja_data.max_column
        lista_fila.append(columna[fila].value)
    datos.append(lista_fila)
secciones = []
for seccion in datos:
    if seccion[0] not in secciones:
        secciones.append(seccion)

###########
#Funciones#
###########

""" ###Funciones graficas### """

def graf_barras(seccion_selec, ventana):
    """
    Función que crea la grafica por la seccion
    """
    ev = ["PEP1","PEP2","Control 1","Control 2"]
    rut = []
    pep1 = []
    pep2 = []
    con1 = []
    con2 = []
    for estudiante in datos: #Para un estudiante de la lista datos
        if estudiante[0] == seccion_selec: #Si el estudiante pertenece a la seccion A-1, se guarda su nota en una lista con la evaluacion correspondiente
            rut.append(estudiante[1])
            pep1.append(float(estudiante[2]))
            pep2.append(float(estudiante[3]))
            con1.append(float(estudiante[4]))
            con2.append(float(estudiante[5]))
    #Se calcula el promedio de la evaluacion por SECCION        
    prom = [(round(stats.mean(pep1), 1)), (round(stats.mean(pep2), 1)), (round(stats.mean(con1), 1)), (round(stats.mean(con2), 1))]
    prom.reverse()
    #Se crea la figura 
    fig, ax = plt.subplots(figsize=(5, 10))
    ax.bar(ev, prom)
    ax.set_ylim(1, 7)
    ax.set_title("Gráfico de Barras " + seccion_selec)
    #Mostrar los valores encima de cada barra
    for i, valor in enumerate(prom):
        ax.text(i, valor + 0.1, str(valor), ha='center', va='bottom', fontsize=10, fontweight='bold')
        # i = posición en X, valor+0.1 = posición Y (arriba de la barra)

    #Se aplica el grafico en Tkinter
    #Se crea una ventana para adjuntar el grafico
    grafico = tk.Toplevel()
    grafico.title("Grafica de Barras")
    #Adjunta el gráfico
    fig.tight_layout()
    canvas = FigureCanvasTkAgg(fig, master=grafico)
    canvas.draw()
    canvas.get_tk_widget().pack(pady=10,fill=tk.BOTH,expand=False)

def graficar_prom_secciones(secciones, ventana): # secciones = ["A-1", "C-3"]
    lista_prom = [] #Se guardan los promedios finales de cada seccion en esta lista vacia
    #Se realiza el procedimiento dos veces para una seccion y para otra
    i = 0
    while i < len(secciones): #Mientras i sea menor al largo de la lista
        pep1 = [] #Se crean listas vacias para guardar las notas
        pep2 = []
        con1 = []
        con2 = []
        seccion_selec = secciones[i]
        for estudiante in datos: #Para un estudiante de la lista datos
            if estudiante[0] == seccion_selec: #Si el estudiante pertenece a la seccion A-1, se guarda su nota en una lista con la evaluacion correspondiente
                pep1.append(float(estudiante[2]))
                pep2.append(float(estudiante[3]))
                con1.append(float(estudiante[4]))
                con2.append(float(estudiante[5]))
        #Se calcula el promedio de la evaluacion por SECCION 
        mejor_control = 1 #Se evalua cual promedio de los controles es mejor
        if (round(stats.mean(con1), 1)) >= (round(stats.mean(con2), 1)): #Si el promedio del control 1 es mayor que el del control 2, se utiliza ese, si no el otro
            mejor_control =  round(stats.mean(con1), 1)
        else:
            mejor_control = round(stats.mean(con2), 1)
        prom = round((round(stats.mean(pep1), 1)*0.4) + (round(stats.mean(pep2), 1)*0.4) + (mejor_control * 0.2), 1) #Se calcula el promedio
        lista_prom.append(prom) #Se guarda el po en la lista
        i+=1
    ventana = tk.Toplevel()
    ventana.title("Grafica de barras")
    #Se crea la figura 
    fig, ax = plt.subplots(figsize=(5, 10))
    ax.bar(secciones, lista_prom)
    ax.set_ylim(1, 7)
    ax.set_title("Gráfico de Barras de las secciones ")
    #Mostrar los valores encima de cada barra
    for i, valor in enumerate(lista_prom):
        ax.text(i, valor + 0.1, str(valor), ha='center', va='bottom', fontsize=10, fontweight='bold')
        # i = posición en X, valor+0.1 = posición Y (arriba de la barra)

    #Se aplica el grafico en Tkinter
    #Se crea una ventana para adjuntar el grafico
    #Adjunta el gráfico
    fig.tight_layout()
    canvas = FigureCanvasTkAgg(fig, master=ventana)
    canvas.draw()
    canvas.get_tk_widget().pack(pady=10,fill=tk.BOTH,expand=False)

def tabla_estudiantes(seccion_selec, ventana):
    """
    Función para obtener una tabla con todos los promedios de los estudiantes de una sección y especular si pueden dar POR o si están reprobados
    """
    datos_seccion = [] #Lista vacia que guardará unicamente los datos de la seccion selecionada
    for seccion in datos: #Para una seccion en datos, si la seccion es igual se guarda 
        if seccion[0] == seccion_selec: 
            datos_seccion.append(seccion)
    datos_estudiante = [] #Lista vacia que guardará los mismos datos menos el peor control
    for estudiante in datos_seccion:
        control_mayor = ""
        if estudiante[4] > estudiante[5]: #Si el control es mayor que el otro, se agrega a una lista
            control_mayor = estudiante[4]
        elif estudiante[5] > estudiante[4]:
            control_mayor = estudiante[5]
        else: #Si ambos controles son iguales, se añade cualquiera de los dos
            control_mayor = estudiante[4]
        estudiante.pop()
        estudiante.pop()
        estudiante.append(str(control_mayor))
        datos_estudiante.append(estudiante)

    datos_tabla = []    #Lista vacia final que se usará para la tabla, donde ademas de los datos conocidos se le agregara el promedio, 
                        #el estado actual del estudiante y si pueda dar POR
    for notas in datos_estudiante:
        prom_estudiante = round((float(notas[2])*0.4) + (float(notas[3])*0.4) + (float(notas[4])*0.2), 1) #Se calcula el promedio
        notas.append(str(prom_estudiante))
        if prom_estudiante >= 4.0: #Si el promedio es mayor o igual a 4 aprueba, si no, reprueba (se agrega esto a la lista)
            notas.append("APROBADO")
        elif 4.0 > prom_estudiante:
            notas.append("REPROBADO")
        if prom_estudiante >= 3.0: #Si el promedio es mayor o igual a 3 puede dar POR, si no, no puede rendir esta prueba (se agrega esto a la lista)
            notas.append("SI")
        elif 3.0 > prom_estudiante:
            notas.append("NO")
        datos_tabla.append(notas)
    encabezados = ["Rut estudiante", "Pep 1", "Pep 2", "Mejor Control", "Promedio", "Estado actual", "¿Habilitado para dar POR?"] #Encabezados para la tabla
    # Ventana nueva
    ventana = tk.Toplevel()
    ventana.title("Tabla "+ seccion_selec)
    # Crear tabla con Treeview modulo de tkinter
    tabla = ttk.Treeview(
    ventana,
    columns=encabezados,
    show="headings"
    )
    # Configurar encabezados
    for col in encabezados:
        tabla.heading(col, text=col) # En que lugar y que dice estara el encabezado
        tabla.column(col, width=130, anchor='center')  # Ancho ajustable            
    # Insertar filas en la tabla
    for fila in datos_tabla:
        tabla.insert("", tk.END, values=fila[1:])  # Saltamos 'C-3' (índice 0) con fila[1:]
    # Agregar scroll si es necesario
    scroll = ttk.Scrollbar(ventana, orient="vertical", command=tabla.yview) # En y
    tabla.configure(yscrollcommand=scroll.set) # Lo instala
    # Posicionar tabla y scroll
    tabla.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    scroll.pack(side=tk.RIGHT, fill=tk.Y)

""" ###Funciones ventanas### """

def cerrar_todo():
    """
    Función que permite cerrar matplotlib _
    """
    root.destroy()
    sys.exit()

def ventana_grafbarras():
    """
    Funcion que genera cada grafica de secciones
    """
    ventana = tk.Toplevel()
    ventana.title("Graficos de barras")
    label = tk.Label(
        ventana,
        text="Por favor, seleccione una sección"
        )
    ventana.geometry("300x400")
    ventana.configure(bg="#5a6879")
    boton_A = tk.Button(
        ventana,
        text="A-1",
        width=10,
        height=2,
        command=lambda: graf_barras("A-1", ventana)
        )
    boton_C = tk.Button(
        ventana,
        text="C-3",
        width=10,
        height=2,
        command=lambda: graf_barras("C-3", ventana)
        )
    label.pack()
    boton_A.pack(pady=70)
    boton_C.pack(pady=70)

def ventana_tablaproms():
    """
    Función que genera cada tabla con los promedios y estado actual de los estudiantes por seccion
    """
    ventana = tk.Toplevel()
    ventana.title("Tabla promedios")
    label = tk.Label(
        ventana,
        text="Por favor, seleccione una sección"
        )
    ventana.geometry("300x400")
    ventana.configure(bg="#5a6879")
    boton_A = tk.Button(
        ventana,
        text="A-1",
        width=10,
        height=2,
        command=lambda: tabla_estudiantes("A-1", ventana)
        )
    boton_C = tk.Button(
        ventana,
        text="C-3",
        width=10,
        height=2,
        command=lambda: tabla_estudiantes("C-3", ventana)
        )
    label.pack()
    boton_A.pack(pady=70)
    boton_C.pack(pady=70)

def ventana_graf_barras2():
    """
    Función que genera un gráfico de barras con los promedios de las secciones
    """
    secciones = ["A-1","C-3"]
    ventana = tk.Toplevel()
    ventana.title("Grafico de barras")
    label = tk.Label(
        ventana,
        text="Por favor, seleccione una sección"
        )
    ventana.geometry("300x400")
    ventana.configure(bg="#5a6879")
    boton_A1 = tk.Button(
        ventana,
        text="Promedio de las secciones",
        height=3,
        command=lambda: graficar_prom_secciones(secciones, ventana)
        )
    label.pack()
    boton_A1.pack(pady=70)

##################
#Bloque principal#
##################

#Ventana principal

root = tk.Tk()
root.title("Notas Smart")

#Saludo Inicial
saludo = tk.Label(
    root,
    text="¡Bienvenido profesor!",
    font=("Arial", 20, "bold"),
    relief="raised",
    bd=10
    )
#Botones
vent_1 = tk.Button(
    root, 
    text = "Promedios por evaluación",
    command = ventana_grafbarras,
    width = 15,
    height = 5,
    padx = 20,
    pady = 10
    )
vent_2 = tk.Button(
    root,
    text = "Tabla con\n promedios de estudiantes",
    command= ventana_tablaproms,
    width=15,
    height=5,
    padx= 20,
    pady=10
    )
vent_3 = tk.Button(
    root,
    text = "Promedios secciones",
    command= ventana_graf_barras2,
    width=15,
    height=5,
    padx= 20,
    pady=10
    )

#Imagen
prin = Image.open("logo-usach.png")
prin = prin.resize((150, 53))
prin_tk = ImageTk.PhotoImage(prin)
et_prin = tk.Label(root, image=prin_tk)
### FUNCIONES NECESARIAS ###
et_prin.pack(pady=30)
saludo.pack(pady=32)
vent_1.pack(padx=10, pady=35)
vent_2.pack(padx=10, pady=35)
vent_3.pack(padx=10, pady=35)
root.geometry("600x700")
root.protocol("WM_DELETE_WINDOW", cerrar_todo)
root.mainloop()
