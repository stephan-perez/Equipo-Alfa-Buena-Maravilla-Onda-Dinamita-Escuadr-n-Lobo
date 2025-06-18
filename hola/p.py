import openpyxl as ox
import tkinter as tk
from matplotlib import *

###obt datos###
datos = []

hoja = ox.load_workbook("nombre.xlsx")

hoja_data = hoja.active

n_datos = 0

for fila in range(1, hoja_data.max_row):
    _fila = []
    n_datos += 1
    for columna in hoja_data.iter_cols(0, hoja_data.max_column):
        _fila.append(columna[fila].value)
    datos.append(_fila)
print(datos)
seccion = []
alumnos = []
notas = []



i = 0
while i < len(datos):
    seccion.append(datos[i][0])
    i+=1
print(seccion)
i = 0
while i < len(datos):
    alumnos.append(datos[i][1])
    i+=1
print(alumnos)
i = 0
while i < len(datos):
    notas.append(datos[i][2])
    i+=1
print(notas)
###obt datos###

