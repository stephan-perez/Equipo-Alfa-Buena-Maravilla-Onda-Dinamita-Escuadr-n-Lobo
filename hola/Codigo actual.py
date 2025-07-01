#Bloque definiciones
#Constantes

#Importaciones
import sys
import tkinter as tk
from PIL import ImageTk, Image
import openpyxl as ox
import matplotlib.pyplot as plt
import statistics as stats
from tkinter import Toplevel
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
#pip install pillow
#Funciones
# obtencion de datos en excel
datos= []
#abre el archivo
hoja = ox.load_workbook("datos_secciones.xlsx")
hoja_data = hoja.active
# de manera ordenada se ingresa los datos en una lista de listas
n_datos = 0
for fila in range(1, hoja_data.max_row):
    lista_fila = []
    n_datos += 1
    for columna in hoja_data.iter_cols(0, hoja_data.max_column):
        lista_fila.append(columna[fila].value)
    datos.append(lista_fila)

""" ###funciones gráficas### """

def cerrar_todo():
    """
    funcion que permite cerrar matplotlib _
    """
    root.destroy()
    sys.exit()

def graf_barrasA(ventana):
    """
    funcion que crea la grafica para la seccion A-1 
    """
    ev = ["PEP1","PEP2","Control 1","Control 2"]
    rut_A1 = []
    pep1_A1 = []
    pep2_A1 = []
    con1_A1 = []
    con2_A1 = []
    # analiza las notas de cada evaluacion y las promedia
    for estudiante in datos:
        if estudiante[0] == "A-1":
            rut_A1.append(estudiante[1])
            pep1_A1.append(float(estudiante[2]))
            pep2_A1.append(float(estudiante[3]))
            con1_A1.append(float(estudiante[4]))
            con2_A1.append(float(estudiante[5]))
            
    prom_A1 = [(round(stats.mean(pep1_A1), 1)), (round(stats.mean(pep2_A1), 1)), (round(stats.mean(con1_A1), 1)), (round(stats.mean(con2_A1), 1))]
    prom_A1.reverse()
    
    # Crear figura------
    fig, ax = plt.subplots(figsize=(5, 10))
    ax.bar(ev, prom_A1)
    ax.set_ylim(1, 7)
    ax.set_title("Gráfico de Barras A-1")
    # Mostrar los valores encima de cada barra
    for i, valor in enumerate(prom_A1):
        ax.text(i, valor + 0.1, str(valor), ha='center', va='bottom', fontsize=10, fontweight='bold')
        # i = posición en X, valor+0.1 = posición Y (arriba de la barra)

    # #grafico en tkinter

    # crea una ventana para adjuntar el gráfico
    grafico_A1 = tk.Toplevel()
    grafico_A1.title("Grafica de Barras")
    
    # adjunta el gráfico
    fig.tight_layout()
    canvas = FigureCanvasTkAgg(fig, master=grafico_A1)
    canvas.draw()
    canvas.get_tk_widget().pack(pady=10,fill=tk.BOTH,expand=False)

def graf_barrasC(ventana):
    """
    funcion que crea la grafica de la seccion C-3
    mismo concepto de la funcion anterior
    """
    ev = ["PEP1","PEP2","Control 1","Control 2"]
    rut_C3 = []
    pep1_C3 = []
    pep2_C3 = []
    con1_C3 = []
    con2_C3 = [] 
    for estudiante in datos: 
        if estudiante[0] == "C-3":
            rut_C3.append(estudiante[1])
            pep1_C3.append(float(estudiante[2]))
            pep2_C3.append(float(estudiante[3]))
            con1_C3.append(float(estudiante[4]))
            con2_C3.append(float(estudiante[5]))
    prom_C3 = [(round(stats.mean(pep1_C3), 1)), (round(stats.mean(pep2_C3), 1)), (round(stats.mean(con1_C3), 1)), (round(stats.mean(con2_C3), 1))]
    prom_C3.reverse()
    
    # Crear figura------
    fig, ax = plt.subplots(figsize=(5, 10))
    ax.bar(ev, prom_C3)
    ax.set_ylim(1, 7)
    ax.set_title("Gráfico de Barra C-3")
    # Mostrar los valores encima de cada barra
    for i, valor in enumerate(prom_C3):
        ax.text(i, valor + 0.1, str(valor), ha='center', va='bottom', fontsize=10, fontweight='bold')
        # i = posición en X, valor+0.1 = posición Y (arriba de la barra)

    # #grafico en tkinter
    grafico_C3 = tk.Toplevel()
    grafico_C3.title("Gráfica")
     # adjunta el grafico a la ventana creada
    fig.tight_layout()
    canvas = FigureCanvasTkAgg(fig, master=grafico_C3)
    canvas.draw()
    canvas.get_tk_widget().pack(pady=10,fill=tk.BOTH,expand=False)

def promediar_pruebas_A1(ventana):
    """
    funcion que promedia cada evaluacion de la seccion A-1
    """
    ev = ["PEP1","PEP2","Control 1","Control 2"]
    prom_pruebas_A1 = []
    i=2
    while i != 6:
        j = 1
        lista_pasajera = []
        while datos[j][0] == "A-1":
            lista_pasajera.append(float(datos[j][i]))
            j+=1
        largo = len(lista_pasajera)
        suma_A1 = sum(lista_pasajera)
        prom_pruebas_A1.append(round((suma_A1/largo),1))
        i+=1
            # Crear figura------
    fig, ax = plt.subplots(figsize=(5, 10))
    ax.bar(ev, prom_pruebas_A1)
    ax.set_ylim(1, 7)
    ax.set_title("caca")
    # Mostrar los valores encima de cada barra
    for i, valor in enumerate(prom_pruebas_A1):
        ax.text(i, valor + 0.1, str(valor), ha='center', va='bottom', fontsize=10, fontweight='bold')
        # i = posición en X, valor+0.1 = posición Y (arriba de la barra)

    # #grafico en tkinter
    # Integrar en Tkinter
    grafico_sec_A1 = tk.Toplevel()
    grafico_sec_A1.title("Gráfica")
    ####
    """
    Acá debería crearse una definición que construya una ventana y en esa ventana
    un boton que crea la grafica..
    """
    fig.tight_layout()
    canvas = FigureCanvasTkAgg(fig, master=grafico_sec_A1)
    canvas.draw()
    canvas.get_tk_widget().pack(pady=10,fill=tk.BOTH,expand=False)

def promediar_pruebas_C3(ventana):
    """
    mismo concepto de la anterior pero en la seccion C-3
    """
    prom_pruebas_C3 = []
    ev = ["PEP1","PEP2","Control 1","Control 2"]
    i=2
    j=1
    while i != 6:
        lista_pasajera = []
        if datos[j][0] == "C-3":
            while j<len(datos) and datos[j][0] == "C-3":
                lista_pasajera.append(float(datos[j][i]))
                j+=1
            j=1
            largo = len(lista_pasajera)
            suma_C3 = sum(lista_pasajera)
            prom_pruebas_C3.append(round((suma_C3/largo),1))
            i+=1
        else:
            j += 1
    # Crear figura------
    fig, ax = plt.subplots(figsize=(5, 10))
    ax.bar(ev, prom_pruebas_C3)
    ax.set_ylim(1, 7)
    ax.set_title("caca")
    # Mostrar los valores encima de cada barra
    for i, valor in enumerate(prom_pruebas_C3):
        ax.text(i, valor + 0.1, str(valor), ha='center', va='bottom', fontsize=10, fontweight='bold')
        # i = posición en X, valor+0.1 = posición Y (arriba de la barra)

    # #grafico en tkinter
    # Integrar en Tkinter
    grafico_sec_C3 = tk.Toplevel()
    grafico_sec_C3.title("Gráfica")
    fig.tight_layout()
    canvas = FigureCanvasTkAgg(fig, master=grafico_sec_C3)
    canvas.draw()
    canvas.get_tk_widget().pack(pady=10,fill=tk.BOTH,expand=False)
    

""" ###VENTANAS### """

def ventana_sec():
    ventana = tk.Toplevel()
    ventana.title("Elija una sección querido profesor.")
    label = tk.Label(
        ventana,
        text= "Elija una sección querido profesor."
    )
    ventana.geometry("300x400")
    ventana.configure(bg="#5a6879")
    boton_A = tk.Button(
        ventana,
        text = "A-1",
        width=10,
        height=2,
        command=lambda: promediar_pruebas_A1(ventana)
        )
    boton_C = tk.Button(
        ventana,
        text = "C-3",
        width=10,
        height=2,
        command=lambda: promediar_pruebas_C3(ventana)
    )
    label.pack()
    boton_A.pack(pady=70)
    boton_C.pack(pady=70)
    
def ventana_barra():
    """
    funcion que genera la segunda ventana con cada grafica de secciones
    """
    ventana = tk.Toplevel()
    ventana.title("Elija un sección querido profesor.")
    label = tk.Label(
        ventana,
        text="Elija un sección querido profesor."
        )
    ventana.geometry("300x400")
    ventana.configure(bg="#5a6879")
    boton_C = tk.Button(
        ventana,
        text="C-3",
        width=10,
        height=2,
        command=lambda: graf_barrasC(ventana)
    )
    boton_A = tk.Button(
        ventana,
        text="A-1",
        width=10,
        height=2,
        command=lambda: graf_barrasA(ventana)
        )
    label.pack()
    boton_A.pack(pady=70)
    boton_C.pack(pady=70)

#Bloque principal
#Ventana principal

root = tk.Tk()
root.title("Notas Smart")

#Saludo
saludo = tk.Label(
    root,
    text="¡Hola profesor!",
    font=("Arial", 20, "bold"),
    relief="raised",
    bd=10
    )
#boton
vent_1 = tk.Button(
    root, 
    text = "Gráfica de barras",
    command = ventana_barra,
    width = 15,
    height = 5,
    padx = 20,
    pady = 10
    )
vent_2 = tk.Button(
    root,
    text = "grafica de barraaaaa",
    command= ventana_sec,
    width=15,
    height=5,
    padx= 20,
    pady=10
)
#Imagen
prin = Image.open("logo-usach.png")
prin = prin.resize((150, 53))
prin_tk = ImageTk.PhotoImage(prin)
et_prin = tk.Label(root, image=prin_tk)
### FUNCIONES NECESARIAS ###
et_prin.pack(pady=30)
saludo.pack(pady=32)
vent_1.pack(padx=10, pady=35)
vent_2.pack(padx=10, pady=35)
root.geometry("600x700")
root.protocol("WM_DELETE_WINDOW", cerrar_todo)
root.mainloop()